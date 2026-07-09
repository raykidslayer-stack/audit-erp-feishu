from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import requests
from tenacity import retry, stop_after_attempt, wait_fixed

from .config import Settings
from .models import CostReconcileItem, CostReconcileSummary


AUDIT_TOKEN_HEADER = "X-Audit-Token"


@retry(stop=stop_after_attempt(3), wait=wait_fixed(5))
def read_cost_reconcile(settings: Settings) -> CostReconcileSummary:
    backend = _read_backend_reconcile(settings)
    if backend:
        return backend

    if not settings.erp_cost_file:
        return CostReconcileSummary(
            report_date=f"{date.today():%Y-%m-%d}",
            source="local",
            status="blocked",
            detail_url=_audit_detail_url(settings),
            error=(
                "缺少 ERP 成本表，无法判断 ERP 缺失、audit 缺失或双方成本不一致。"
                "请配置 ERP_COST_FILE，或让 audit 后端提供 /api/cost_reconcile。"
            ),
        )

    audit_rows = _read_audit_costs(settings)
    erp_rows = _read_erp_cost_file(Path(settings.erp_cost_file))
    return _reconcile_costs(audit_rows, erp_rows, settings)


def _read_backend_reconcile(settings: Settings) -> CostReconcileSummary | None:
    session = _audit_session(settings)
    response = session.get(_audit_api_url(settings, "/api/cost_reconcile"), timeout=30)
    if response.status_code == 404:
        return None
    response.raise_for_status()
    data = response.json()
    if data.get("status") not in {"success", "ok"}:
        raise RuntimeError(f"audit cost reconcile failed: {data}")
    return _summary_from_backend(data, settings)


def _read_audit_costs(settings: Settings) -> list[dict[str, Any]]:
    session = _audit_session(settings)
    response = session.get(_audit_api_url(settings, "/api/cost_validate"), timeout=30)
    response.raise_for_status()
    data = response.json()
    if data.get("status") not in {"success", "ok"}:
        raise RuntimeError(f"audit cost validate failed: {data}")
    return [
        {
            "name": item.get("sku", item.get("name", "")),
            "cost": item.get("raw_cost", item.get("cost", "")),
            "source_row": item.get("row", ""),
        }
        for item in data.get("items", [])
    ]


def _audit_session(settings: Settings) -> requests.Session:
    session = requests.Session()
    if settings.audit_token:
        session.headers.update({AUDIT_TOKEN_HEADER: settings.audit_token})
        return session

    login_url = _audit_api_url(settings, "/api/login")
    login = session.post(
        login_url,
        data={"username": settings.audit_username, "password": settings.audit_password},
        timeout=20,
    )
    login.raise_for_status()
    data = login.json()
    if data.get("status") != "success":
        raise RuntimeError(f"audit login failed: {data}")
    return session


def _audit_api_url(settings: Settings, path: str) -> str:
    base = settings.audit_url.split("/daily")[0].rstrip("/")
    return f"{base}{path}"


def _audit_detail_url(settings: Settings) -> str:
    base = settings.audit_url.split("/daily")[0].rstrip("/")
    return f"{base}/costs"


def _read_erp_cost_file(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise RuntimeError(f"ERP cost file not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_erp_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_erp_xlsx(path)
    raise RuntimeError(f"Unsupported ERP cost file type: {path.suffix}")


def _read_erp_csv(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return _rows_from_table(list(csv.reader(handle)))
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Unable to decode ERP cost CSV: {path}")


def _read_erp_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read ERP .xlsx cost files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if getattr(sheet, "max_column", 0) <= 1 and hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    table = [[cell for cell in row] for row in sheet.iter_rows(max_col=10, values_only=True)]
    return _rows_from_table(table)


def _rows_from_table(table: list[list[Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not table:
        return rows

    headers = [str(value or "").strip() for value in table[0]]
    name_idx = _find_column(headers, ["货品名称", "商品名称", "品名", "货品名"], default=2)
    cost_idx = _find_column(headers, ["成本价", "成本价（￥）", "成本价(￥)", "成本"], default=4)

    for row_number, row in enumerate(table[1:], start=2):
        name = _cell(row, name_idx)
        cost = _cell(row, cost_idx)
        if not name and not cost:
            continue
        rows.append({"name": name, "cost": cost, "source_row": row_number})
    return rows


def _find_column(headers: list[str], names: list[str], default: int) -> int:
    normalized = {_normalize_header(header): idx for idx, header in enumerate(headers)}
    for name in names:
        idx = normalized.get(_normalize_header(name))
        if idx is not None:
            return idx
    return default


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", value)).lower()


def _cell(row: list[Any], idx: int) -> str:
    if idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def _reconcile_costs(
    audit_rows: list[dict[str, Any]],
    erp_rows: list[dict[str, Any]],
    settings: Settings,
) -> CostReconcileSummary:
    audit_by_key = _index_rows(audit_rows)
    erp_by_key = _index_rows(erp_rows)
    items: list[CostReconcileItem] = []

    for key in sorted(set(erp_by_key) - set(audit_by_key)):
        row = erp_by_key[key][0]
        items.append(
            CostReconcileItem(
                issue_type="audit_missing",
                action="去 audit 成本库补品名和成本",
                erp_name=row["name"],
                erp_cost=row["cost"],
                note="ERP 有该品名，audit 成本库没有。",
            )
        )
        if _missing_cost(_decimal_or_none(row["cost"])):
            items.append(
                CostReconcileItem(
                    issue_type="erp_cost_missing",
                    action="去 ERP 补成本",
                    erp_name=row["name"],
                    erp_cost=row["cost"],
                    note="ERP 有该品名但成本为空、0 或负数；audit 也缺该品名。",
                )
            )

    for key in sorted(set(audit_by_key) - set(erp_by_key)):
        row = audit_by_key[key][0]
        items.append(
            CostReconcileItem(
                issue_type="erp_missing",
                action="确认是否停用；仍在售则去 ERP 补商品资料",
                audit_name=row["name"],
                audit_cost=row["cost"],
                note="audit 有该品名，ERP 成本表没有。",
            )
        )
        if _missing_cost(_decimal_or_none(row["cost"])):
            items.append(
                CostReconcileItem(
                    issue_type="audit_cost_missing",
                    action="去 audit 成本库补成本",
                    audit_name=row["name"],
                    audit_cost=row["cost"],
                    note="audit 有该品名但成本为空、0 或负数；ERP 也缺该品名。",
                )
            )

    for key in sorted(set(audit_by_key) & set(erp_by_key)):
        audit_group = audit_by_key[key]
        erp_group = erp_by_key[key]
        if len(audit_group) > 1 or len(erp_group) > 1:
            items.append(
                CostReconcileItem(
                    issue_type="duplicate_name",
                    action="先去对应系统合并或区分重复品名",
                    erp_name=erp_group[0]["name"],
                    audit_name=audit_group[0]["name"],
                    note=f"ERP 重复 {len(erp_group)} 条，audit 重复 {len(audit_group)} 条。",
                )
            )
            continue

        audit_row = audit_group[0]
        erp_row = erp_group[0]
        audit_cost = _decimal_or_none(audit_row["cost"])
        erp_cost = _decimal_or_none(erp_row["cost"])
        if not _same_visible_name(audit_row["name"], erp_row["name"]):
            items.append(
                CostReconcileItem(
                    issue_type="name_mismatch",
                    action="统一品名或维护品名映射",
                    erp_name=erp_row["name"],
                    audit_name=audit_row["name"],
                    erp_cost=erp_row["cost"],
                    audit_cost=audit_row["cost"],
                    note="标准化后能匹配，但原始品名不完全一致。",
                )
            )
        if _missing_cost(audit_cost):
            items.append(
                CostReconcileItem(
                    issue_type="audit_cost_missing",
                    action="去 audit 成本库补成本",
                    erp_name=erp_row["name"],
                    audit_name=audit_row["name"],
                    erp_cost=erp_row["cost"],
                    audit_cost=audit_row["cost"],
                    note="品名已对上，audit 成本为空、0 或负数。",
                )
            )
        if _missing_cost(erp_cost):
            items.append(
                CostReconcileItem(
                    issue_type="erp_cost_missing",
                    action="去 ERP 补成本",
                    erp_name=erp_row["name"],
                    audit_name=audit_row["name"],
                    erp_cost=erp_row["cost"],
                    audit_cost=audit_row["cost"],
                    note="品名已对上，ERP 成本为空、0 或负数。",
                )
            )
        if audit_cost is not None and erp_cost is not None and audit_cost > 0 and erp_cost > 0:
            diff = audit_cost - erp_cost
            if diff.copy_abs() >= Decimal("0.01"):
                items.append(
                    CostReconcileItem(
                        issue_type="cost_mismatch",
                        action="财务确认基准价后修正对应系统",
                        erp_name=erp_row["name"],
                        audit_name=audit_row["name"],
                        erp_cost=_money(erp_cost),
                        audit_cost=_money(audit_cost),
                        diff=_money(diff),
                        note="品名已对上，但两边成本不一致。",
                    )
                )

    counts = Counter(item.issue_type for item in items)
    erp_file = Path(settings.erp_cost_file) if settings.erp_cost_file else None
    stale = False
    if erp_file and erp_file.exists():
        modified = datetime.fromtimestamp(erp_file.stat().st_mtime, timezone.utc)
        stale = (datetime.now(timezone.utc) - modified).days > 30

    return CostReconcileSummary(
        report_date=f"{date.today():%Y-%m-%d}",
        source="local",
        status="success",
        audit_count=len(audit_rows),
        erp_count=len(erp_rows),
        total_issues=len(items),
        audit_missing_count=counts["audit_missing"],
        erp_missing_count=counts["erp_missing"],
        name_mismatch_count=counts["name_mismatch"],
        audit_cost_missing_count=counts["audit_cost_missing"],
        erp_cost_missing_count=counts["erp_cost_missing"],
        cost_mismatch_count=counts["cost_mismatch"],
        duplicate_count=counts["duplicate_name"],
        stale_erp_file=stale,
        detail_url=_audit_detail_url(settings),
        items=items,
    )


def _summary_from_backend(data: dict[str, Any], settings: Settings) -> CostReconcileSummary:
    items = [
        CostReconcileItem(
            issue_type=str(item.get("issue_type", "")),
            action=str(item.get("action", "")),
            erp_name=str(item.get("erp_name", "")),
            audit_name=str(item.get("audit_name", "")),
            erp_cost=str(item.get("erp_cost", "")),
            audit_cost=str(item.get("audit_cost", "")),
            diff=str(item.get("diff", "")),
            note=str(item.get("note", "")),
        )
        for item in data.get("items", [])
    ]
    return CostReconcileSummary(
        report_date=str(data.get("report_date") or f"{date.today():%Y-%m-%d}"),
        source="audit",
        status="success",
        audit_count=int(data.get("audit_count", 0) or 0),
        erp_count=int(data.get("erp_count", 0) or 0),
        total_issues=int(data.get("total_issues", len(items)) or 0),
        audit_missing_count=int(data.get("audit_missing_count", 0) or 0),
        erp_missing_count=int(data.get("erp_missing_count", 0) or 0),
        name_mismatch_count=int(data.get("name_mismatch_count", 0) or 0),
        audit_cost_missing_count=int(data.get("audit_cost_missing_count", 0) or 0),
        erp_cost_missing_count=int(data.get("erp_cost_missing_count", 0) or 0),
        cost_mismatch_count=int(data.get("cost_mismatch_count", 0) or 0),
        duplicate_count=int(data.get("duplicate_count", 0) or 0),
        stale_erp_file=bool(data.get("stale_erp_file", False)),
        detail_url=str(data.get("detail_url") or _audit_detail_url(settings)),
        items=items,
    )


def _index_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    indexed: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        name = str(row.get("name", "")).strip()
        key = normalize_product_name(name)
        if not key:
            continue
        indexed.setdefault(key, []).append(
            {"name": name, "cost": str(row.get("cost", "")).strip()}
        )
    return indexed


def normalize_product_name(name: str) -> str:
    value = unicodedata.normalize("NFKC", name or "")
    value = value.replace("（", "(").replace("）", ")")
    value = re.sub(r"\s+", "", value)
    return value.lower()


def _same_visible_name(left: str, right: str) -> bool:
    return unicodedata.normalize("NFKC", left).strip() == unicodedata.normalize("NFKC", right).strip()


def _decimal_or_none(value: Any) -> Decimal | None:
    text = str(value or "").strip().replace(",", "").replace("￥", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _missing_cost(value: Decimal | None) -> bool:
    return value is None or value <= 0


def _money(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'))}"
